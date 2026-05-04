"""Extrae TODOS los comments (notas) de cada celda de la hoja 'Input sheet'
del Excel fcffsimpleginzu.xlsx para entender exactamente qué hace cada input
de Damodaran (especialmente 'Years since last 10K').
"""
import sys
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FP = ROOT / "data" / "raw_xbrl" / "fcffsimpleginzu.xlsx"
if not FP.exists():
    # fallback Downloads
    FP = Path.home() / "Downloads" / "fcffsimpleginzu.xlsx"
if not FP.exists():
    print(f"No encuentro fcffsimpleginzu.xlsx")
    sys.exit(1)

print(f"Leyendo {FP}\n")
wb = load_workbook(FP, data_only=False)
print(f"Hojas: {wb.sheetnames}\n")

# Buscar la hoja 'Input sheet' (case-insensitive)
target = None
for n in wb.sheetnames:
    if "input" in n.lower():
        target = n
        break
if target is None:
    target = wb.sheetnames[0]
print(f"=== Hoja: {target} ===\n")

ws = wb[target]
print(f"Dimensions: {ws.dimensions}\n")

# Recorrer celdas y mostrar (cell, value, comment)
print(f"{'Celda':<6} {'Valor':<35} | Notas (comments)")
print("-" * 100)
for row in ws.iter_rows(min_row=1, max_row=80, max_col=10):
    for cell in row:
        v = cell.value
        c = cell.comment
        if v is None and c is None:
            continue
        v_str = str(v)[:32] if v is not None else ""
        c_str = ""
        if c is not None:
            c_str = c.text.replace("\n", " | ")[:500]
        print(f"{cell.coordinate:<6} {v_str:<35} | {c_str}")

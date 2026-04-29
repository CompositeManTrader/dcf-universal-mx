"""
Exporta el DCF a Excel con FORMULAS reales (no valores estaticos).

Esto permite al analista:
  1. Auditar las formulas celda por celda
  2. Editar inputs y ver cambios
  3. Replicar el modelo en Excel sin Python

Layout:
  Sheet 'Inputs'         - Drivers editables (revenue growth, margin, tax, WACC)
  Sheet 'Projection'     - Proyeccion 10y con formulas
  Sheet 'Bridge'         - EV -> Equity Value -> Value/share
  Sheet 'Audit'          - Comparacion vs el calculo Python (para QA)
"""
from __future__ import annotations

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# Damodaran-style colors
GREEN_LIGHT = PatternFill("solid", fgColor="DCEDC8")     # input cells
GREEN_INPUT = PatternFill("solid", fgColor="C5E1A5")     # editable
GREY_BG = PatternFill("solid", fgColor="F1F8E9")         # calculated
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")     # bloomberg blue
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN = Side(border_style="thin", color="888888")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_header(ws, row, cols, label_widths):
    for i, (label, width) in enumerate(zip(cols, label_widths)):
        c = ws.cell(row=row, column=i+1, value=label)
        c.fill = HEADER_FILL
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
        ws.column_dimensions[get_column_letter(i+1)].width = width


def export_dcf_to_excel(base, assumptions, output) -> bytes:
    """
    Genera un Excel binario con formulas DCF.
    Returns bytes para usar con st.download_button.
    """
    wb = Workbook()
    n = assumptions.forecast_years

    # =====================================================================
    # SHEET 1: INPUTS (editable cells in green)
    # =====================================================================
    ws_in = wb.active
    ws_in.title = "Inputs"

    # Empresa
    ws_in["A1"] = "DCF Inputs"
    ws_in["A1"].font = Font(bold=True, size=14, color="1F4E79")

    ws_in["A3"] = "Ticker"
    ws_in["B3"] = base.ticker
    ws_in["A4"] = "Currency"
    ws_in["B4"] = "MDP (millones de pesos)"

    # Base year financials
    row = 6
    ws_in.cell(row=row, column=1, value="Base year financials").font = BOLD_FONT
    base_data = [
        ("Revenue (12M)",        base.revenue),
        ("EBIT (12M)",           base.ebit),
        ("Interest expense",     base.interest_expense),
        ("Cash",                 base.cash),
        ("Total Debt + Lease",   base.financial_debt),
        ("Minority interest",    base.minority_interest),
        ("Non-operating assets", base.non_operating_assets),
        ("Equity book value",    base.equity_book),
        ("Invested Capital",     base.invested_capital),
        ("Shares outstanding",   base.shares_outstanding),
        ("Effective tax rate",   base.effective_tax_rate),
    ]
    for label, val in base_data:
        row += 1
        ws_in.cell(row=row, column=1, value=label)
        c = ws_in.cell(row=row, column=2, value=val)
        c.fill = GREY_BG
        c.number_format = "#,##0.00" if abs(val) > 100 else "0.0000"

    # Drivers (editable)
    row += 2
    ws_in.cell(row=row, column=1, value="Drivers (editable - cambia y ve impact)").font = BOLD_FONT
    drivers = [
        ("Revenue growth (Y1-5)",      assumptions.revenue_growth_high),
        ("Terminal growth",            assumptions.terminal_growth),
        ("Target op margin (Y10)",     assumptions.target_op_margin),
        ("Sales-to-Capital",           assumptions.sales_to_capital),
        ("Marginal tax rate (Y10)",    assumptions.marginal_tax_terminal),
        ("Risk-free rate",             assumptions.risk_free),
        ("ERP",                        assumptions.erp),
        ("Unlevered beta",             assumptions.unlevered_beta),
        ("Terminal WACC",              assumptions.terminal_wacc_override or 0.085),
        ("Initial WACC (computed)",    output.wacc_result.wacc),
        ("Market price",               assumptions.market_price or 0),
    ]
    driver_refs = {}  # name -> cell ref (e.g., "B14")
    for label, val in drivers:
        row += 1
        ws_in.cell(row=row, column=1, value=label)
        c = ws_in.cell(row=row, column=2, value=val)
        c.fill = GREEN_INPUT
        c.font = BOLD_FONT
        c.number_format = "0.0000"
        c.border = BOX
        driver_refs[label] = f"Inputs!$B${row}"

    # =====================================================================
    # SHEET 2: PROJECTION (formulas)
    # =====================================================================
    ws_p = wb.create_sheet("Projection")
    year_cols = ["Concepto"] + [f"Y{i}" for i in range(1, n+1)] + ["Terminal"]
    _set_header(ws_p, 1, year_cols, [28] + [12] * (n+1))

    # Helper to write a row
    def _write_row(row_idx, label, base_val, formula_func, fmt="#,##0.00"):
        """formula_func(year_idx) returns the formula string for year `year_idx` (1-based)."""
        ws_p.cell(row=row_idx, column=1, value=label).font = BOLD_FONT
        ws_p.cell(row=row_idx, column=1).fill = GREY_BG
        for t in range(1, n+1):
            c = ws_p.cell(row=row_idx, column=t+1, value=formula_func(t))
            c.number_format = fmt
            c.fill = GREEN_LIGHT if label in ("Revenue", "EBIT", "FCFF", "PV(FCFF)") else None
        # Terminal column
        c = ws_p.cell(row=row_idx, column=n+2, value=formula_func("term"))
        c.number_format = fmt
        c.fill = GREEN_LIGHT if label in ("Revenue", "EBIT", "FCFF", "PV(FCFF)") else None
        return row_idx

    # Cell references for cross-sheet formulas
    rev_g_ref = driver_refs["Revenue growth (Y1-5)"]
    term_g_ref = driver_refs["Terminal growth"]
    margin_ref = driver_refs["Target op margin (Y10)"]
    s2c_ref = driver_refs["Sales-to-Capital"]
    tax_ref = driver_refs["Marginal tax rate (Y10)"]
    init_wacc_ref = driver_refs["Initial WACC (computed)"]
    term_wacc_ref = driver_refs["Terminal WACC"]
    base_rev_ref = "Inputs!$B$7"
    base_ebit_ref = "Inputs!$B$8"
    base_etr_ref = "Inputs!$B$17"

    # Pre-compute static fade values from output (since Excel formulas for fade are complex)
    # We use the output values directly as the formulas would get unwieldy.
    # The key inputs (drivers) ARE editable; intermediate fades are computed.

    # Row 2: Revenue
    ws_p.cell(row=2, column=1, value="Revenue").font = BOLD_FONT
    ws_p.cell(row=2, column=1).fill = GREEN_LIGHT
    # Y1 = base_rev * (1 + growth_y1)
    # Use output values for growth per year (already computed)
    high_n = assumptions.high_growth_years
    for t in range(1, n+1):
        if t <= high_n:
            growth_formula = rev_g_ref
        else:
            # Linear fade
            step = t - high_n
            steps_remaining = n - high_n
            # g = high + (term - high) * step / steps_remaining
            growth_formula = f"({rev_g_ref}+({term_g_ref}-{rev_g_ref})*{step}/{steps_remaining})"
        if t == 1:
            ws_p.cell(row=2, column=t+1, value=f"={base_rev_ref}*(1+{growth_formula})")
        else:
            prev_col = get_column_letter(t)
            ws_p.cell(row=2, column=t+1, value=f"={prev_col}2*(1+{growth_formula})")
        ws_p.cell(row=2, column=t+1).number_format = "#,##0.00"
    # Terminal
    ws_p.cell(row=2, column=n+2, value=f"={get_column_letter(n+1)}2*(1+{term_g_ref})")
    ws_p.cell(row=2, column=n+2).number_format = "#,##0.00"
    ws_p.cell(row=2, column=n+2).fill = GREEN_LIGHT

    # Row 3: Revenue growth (display)
    ws_p.cell(row=3, column=1, value="Revenue growth")
    for t in range(1, n+1):
        col = get_column_letter(t+1)
        prev_col = get_column_letter(t) if t > 1 else None
        if t == 1:
            ws_p.cell(row=3, column=t+1, value=f"={col}2/{base_rev_ref}-1")
        else:
            ws_p.cell(row=3, column=t+1, value=f"={col}2/{prev_col}2-1")
        ws_p.cell(row=3, column=t+1).number_format = "0.00%"
    ws_p.cell(row=3, column=n+2, value=f"={term_g_ref}")
    ws_p.cell(row=3, column=n+2).number_format = "0.00%"

    # Row 4: Op Margin (linear interpolation base->target in Y10)
    ws_p.cell(row=4, column=1, value="Op Margin")
    base_margin_val = base.ebit / base.revenue if base.revenue else 0
    for t in range(1, n+1):
        # margin_t = base + (target - base) * t/n
        formula = f"({base_margin_val}+({margin_ref}-{base_margin_val})*{t}/{n})"
        ws_p.cell(row=4, column=t+1, value=f"={formula}")
        ws_p.cell(row=4, column=t+1).number_format = "0.00%"
    ws_p.cell(row=4, column=n+2, value=f"={margin_ref}")
    ws_p.cell(row=4, column=n+2).number_format = "0.00%"

    # Row 5: EBIT = Revenue × Margin
    ws_p.cell(row=5, column=1, value="EBIT").font = BOLD_FONT
    ws_p.cell(row=5, column=1).fill = GREEN_LIGHT
    for t in range(1, n+2):
        col = get_column_letter(t+1)
        ws_p.cell(row=5, column=t+1, value=f"={col}2*{col}4")
        ws_p.cell(row=5, column=t+1).number_format = "#,##0.00"
        ws_p.cell(row=5, column=t+1).fill = GREEN_LIGHT

    # Row 6: Tax rate (linear base -> marginal)
    ws_p.cell(row=6, column=1, value="Tax rate")
    for t in range(1, n+1):
        formula = f"({base_etr_ref}+({tax_ref}-{base_etr_ref})*{t}/{n})"
        ws_p.cell(row=6, column=t+1, value=f"={formula}")
        ws_p.cell(row=6, column=t+1).number_format = "0.00%"
    ws_p.cell(row=6, column=n+2, value=f"={tax_ref}")
    ws_p.cell(row=6, column=n+2).number_format = "0.00%"

    # Row 7: NOPAT = EBIT * (1-tax)
    ws_p.cell(row=7, column=1, value="NOPAT (EBIT × (1-t))")
    for t in range(1, n+2):
        col = get_column_letter(t+1)
        ws_p.cell(row=7, column=t+1, value=f"={col}5*(1-{col}6)")
        ws_p.cell(row=7, column=t+1).number_format = "#,##0.00"

    # Row 8: ΔRevenue
    ws_p.cell(row=8, column=1, value="Δ Revenue")
    for t in range(1, n+1):
        col = get_column_letter(t+1)
        if t == 1:
            ws_p.cell(row=8, column=t+1, value=f"={col}2-{base_rev_ref}")
        else:
            prev = get_column_letter(t)
            ws_p.cell(row=8, column=t+1, value=f"={col}2-{prev}2")
        ws_p.cell(row=8, column=t+1).number_format = "#,##0.00"
    ws_p.cell(row=8, column=n+2, value=f"={get_column_letter(n+2)}2-{get_column_letter(n+1)}2")
    ws_p.cell(row=8, column=n+2).number_format = "#,##0.00"

    # Row 9: Reinvestment = ΔRev / S2C
    ws_p.cell(row=9, column=1, value="Reinvestment")
    for t in range(1, n+2):
        col = get_column_letter(t+1)
        ws_p.cell(row=9, column=t+1, value=f"={col}8/{s2c_ref}")
        ws_p.cell(row=9, column=t+1).number_format = "#,##0.00"

    # Row 10: FCFF = NOPAT - Reinvestment
    ws_p.cell(row=10, column=1, value="FCFF").font = BOLD_FONT
    ws_p.cell(row=10, column=1).fill = GREEN_LIGHT
    for t in range(1, n+2):
        col = get_column_letter(t+1)
        ws_p.cell(row=10, column=t+1, value=f"={col}7-{col}9")
        ws_p.cell(row=10, column=t+1).number_format = "#,##0.00"
        ws_p.cell(row=10, column=t+1).fill = GREEN_LIGHT

    # Row 11: WACC (linear initial -> terminal)
    ws_p.cell(row=11, column=1, value="WACC")
    for t in range(1, n+1):
        formula = f"({init_wacc_ref}+({term_wacc_ref}-{init_wacc_ref})*{t}/{n})"
        ws_p.cell(row=11, column=t+1, value=f"={formula}")
        ws_p.cell(row=11, column=t+1).number_format = "0.00%"
    ws_p.cell(row=11, column=n+2, value=f"={term_wacc_ref}")
    ws_p.cell(row=11, column=n+2).number_format = "0.00%"

    # Row 12: Discount factor (cumulative)
    ws_p.cell(row=12, column=1, value="Discount factor")
    for t in range(1, n+1):
        col = get_column_letter(t+1)
        if t == 1:
            ws_p.cell(row=12, column=t+1, value=f"=1/(1+{col}11)")
        else:
            prev = get_column_letter(t)
            ws_p.cell(row=12, column=t+1, value=f"={prev}12/(1+{col}11)")
        ws_p.cell(row=12, column=t+1).number_format = "0.0000"

    # Row 13: PV(FCFF)
    ws_p.cell(row=13, column=1, value="PV(FCFF)").font = BOLD_FONT
    ws_p.cell(row=13, column=1).fill = GREEN_LIGHT
    for t in range(1, n+1):
        col = get_column_letter(t+1)
        ws_p.cell(row=13, column=t+1, value=f"={col}10*{col}12")
        ws_p.cell(row=13, column=t+1).number_format = "#,##0.00"
        ws_p.cell(row=13, column=t+1).fill = GREEN_LIGHT

    # =====================================================================
    # SHEET 3: BRIDGE (EV -> Equity)
    # =====================================================================
    ws_b = wb.create_sheet("Bridge")
    ws_b["A1"] = "Valuation Bridge"
    ws_b["A1"].font = Font(bold=True, size=14, color="1F4E79")

    last_col = get_column_letter(n+1)  # Y10 column
    term_col = get_column_letter(n+2)  # Terminal column

    bridge_rows = [
        ("Sum PV FCFF (Y1-Y10)",     f"=SUM(Projection!B13:{last_col}13)"),
        ("Terminal FCFF (Y11)",       f"=Projection!{term_col}10"),
        ("Terminal WACC",              f"={term_wacc_ref}"),
        ("Terminal growth",            f"={term_g_ref}"),
        ("Terminal Value",             f"=Projection!{term_col}10/({term_wacc_ref}-{term_g_ref})"),
        ("PV(Terminal Value)",         f"=B7*Projection!{last_col}12"),
        ("Enterprise Value",           "=B3+B8"),
        ("(-) Total Debt",             f"=Inputs!$B$11"),
        ("(+) Cash",                   f"=Inputs!$B$10"),
        ("(-) Minority Interest",      f"=Inputs!$B$12"),
        ("(+) Non-operating Assets",   f"=Inputs!$B$13"),
        ("Equity Value",               "=B9-B10+B11-B12+B13"),
        ("Shares (mn)",                f"=Inputs!$B$15/1000000"),
        ("Value per Share (MXN)",      "=B14/B15"),
        ("Market Price (MXN)",         f"={driver_refs['Market price']}"),
        ("Upside / (Downside) %",      "=B16/B17-1"),
    ]
    for i, (label, formula) in enumerate(bridge_rows):
        r = i + 3
        ws_b.cell(row=r, column=1, value=label)
        c = ws_b.cell(row=r, column=2, value=formula)
        if label in ("Enterprise Value", "Equity Value", "Value per Share (MXN)", "Upside / (Downside) %"):
            c.font = WHITE_FONT
            c.fill = HEADER_FILL
            ws_b.cell(row=r, column=1).font = BOLD_FONT
        if "%" in label or label.startswith("Terminal "):
            c.number_format = "0.00%"
        else:
            c.number_format = "#,##0.00"
    ws_b.column_dimensions["A"].width = 32
    ws_b.column_dimensions["B"].width = 18

    # =====================================================================
    # SHEET 4: AUDIT (vs Python computation)
    # =====================================================================
    ws_a = wb.create_sheet("Audit")
    ws_a["A1"] = "Audit: Excel formulas vs Python output"
    ws_a["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws_a["A2"] = "Si las celdas verdes muestran 0 o muy chico, Excel y Python coinciden."

    audit_rows = [
        ("Concepto",              "Python output",                  "Excel formula",        "Diff"),
        ("Sum PV FCFF",           output.sum_pv_fcff,               "=Bridge!B3",            "=B4-C4"),
        ("Terminal Value",        output.terminal_value,             "=Bridge!B7",            "=B5-C5"),
        ("PV Terminal",           output.pv_terminal,                "=Bridge!B8",            "=B6-C6"),
        ("Enterprise Value",      output.enterprise_value,           "=Bridge!B9",            "=B7-C7"),
        ("Equity Value",          output.equity_value,               "=Bridge!B14",           "=B8-C8"),
        ("Value per Share",       output.value_per_share,            "=Bridge!B16",           "=B9-C9"),
    ]
    for i, row in enumerate(audit_rows):
        r = i + 4
        for j, val in enumerate(row):
            c = ws_a.cell(row=r, column=j+1, value=val)
            if i == 0:
                c.font = WHITE_FONT
                c.fill = HEADER_FILL
            elif j == 3:
                c.fill = GREEN_LIGHT
                c.number_format = "#,##0.0000"
            elif j > 0:
                c.number_format = "#,##0.00"
    for i, w in enumerate([28, 18, 18, 18]):
        ws_a.column_dimensions[get_column_letter(i+1)].width = w

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
